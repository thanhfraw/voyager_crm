from fastapi import APIRouter, UploadFile, File, Depends, HTTPException
from openpyxl import load_workbook
from io import BytesIO
from config import database
from models import customers, customer_history, import_logs, nationalities, customer_types, industries, enterprise_types
from auth import require_editor, TokenData

router = APIRouter(prefix="/api/import", tags=["import"])

COLUMN_MAP = {
    "CustomerID":        "customer_id",
    "ShortName":         "short_name",
    "CustKeyCode":       "cust_key_code",
    "CustomerNameEN":    "name_en",
    "CustomerNameVN":    "name_vn",
    "TaxCode":           "tax_code",
    "CompanyPhone":      "phone",
    "CompanyEmail":      "email",
    "Status":            "status",
    # Lookup fields resolved by name
    "CustomerTypeID":    "_customer_type",
    "EnterpriseTypeID":  "_enterprise_type",
    "IndustryID":        "_industry",
    "NationalityID":     "_nationality",
}

async def get_lookup_map(table, name_col="name"):
    rows = await database.fetch_all(table.select())
    return {r["name"].strip().lower(): r["id"] for r in rows}

@router.post("/excel")
async def import_excel(
    file: UploadFile = File(...),
    current_user: TokenData = Depends(require_editor),
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx or .xls files accepted")

    contents = await file.read()
    wb = load_workbook(BytesIO(contents), read_only=True, data_only=True)
    ws = wb.active

    headers = [str(cell.value).strip() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    # Preload lookup maps
    ct_map   = await get_lookup_map(customer_types)
    et_map   = await get_lookup_map(enterprise_types)
    ind_map  = await get_lookup_map(industries)
    nat_map  = await get_lookup_map(nationalities)

    success, errors = 0, []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_data = {headers[i]: (row[i] or "") for i in range(len(headers)) if i < len(row)}

        customer_id = str(row_data.get("CustomerID", "")).strip()
        if not customer_id:
            errors.append({"row": row_num, "error": "Missing CustomerID"})
            continue

        record = {}
        for col, field in COLUMN_MAP.items():
            val = str(row_data.get(col, "")).strip()
            if field.startswith("_"):
                # Resolve lookup
                vl = val.lower()
                if field == "_customer_type":   record["customer_type_id"]   = ct_map.get(vl)
                elif field == "_enterprise_type": record["enterprise_type_id"] = et_map.get(vl)
                elif field == "_industry":       record["industry_id"]        = ind_map.get(vl)
                elif field == "_nationality":    record["nationality_id"]     = nat_map.get(vl)
            else:
                record[field] = val or None

        try:
            existing = await database.fetch_one(
                customers.select().where(customers.c.customer_id == customer_id)
            )
            if existing:
                old = dict(existing._mapping)
                await database.execute(
                    customers.update().where(customers.c.customer_id == customer_id).values(**record)
                )
                action = "UPDATE"
            else:
                await database.execute(customers.insert().values(**record))
                old = None
                action = "CREATE"

            await database.execute(customer_history.insert().values(
                customer_id=customer_id,
                action=action,
                changed_by=current_user.user_id,
                old_data=old,
                new_data=record,
            ))
            success += 1
        except Exception as e:
            errors.append({"row": row_num, "customer_id": customer_id, "error": str(e)})

    log_id = await database.execute(import_logs.insert().values(
        filename=file.filename,
        imported_by=current_user.user_id,
        row_count=row_num - 1,
        success_count=success,
        error_count=len(errors),
        errors=errors[:50] if errors else None,
    ))

    return {
        "log_id": log_id,
        "filename": file.filename,
        "total_rows": row_num - 1,
        "success": success,
        "errors": len(errors),
        "error_details": errors[:20],
    }

@router.get("/logs")
async def get_import_logs(current_user: TokenData = Depends(require_editor)):
    query = """
        SELECT il.*, u.username
        FROM import_logs il
        LEFT JOIN users u ON il.imported_by = u.id
        ORDER BY il.created_at DESC
        LIMIT 50
    """
    rows = await database.fetch_all(query)
    return [dict(r._mapping) for r in rows]
